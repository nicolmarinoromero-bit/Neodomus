"""
Generación de reportes PDF y Excel para NEODOMUS.

PDF: usa reportlab (misma paleta y estilos que factura_service.py).
Excel: usa openpyxl.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ── Paleta NEODOMUS ───────────────────────────────────────────────

ORO = colors.HexColor("#caa24d")
ORO_CLARO = colors.HexColor("#f0c96f")
NEGRO = colors.HexColor("#000000")
FONDO_TABLA = colors.HexColor("#f7f3ea")
GRIS = colors.HexColor("#6b6b6b")
BLANCO = colors.white

# ── Estilos PDF ───────────────────────────────────────────────────


def _styles():
    base = getSampleStyleSheet()
    return {
        "titulo": ParagraphStyle(
            "titulo", parent=base["Title"], fontName="Helvetica-Bold",
            fontSize=18, textColor=ORO, alignment=TA_CENTER, spaceAfter=2,
        ),
        "sub": ParagraphStyle(
            "sub", fontName="Helvetica", fontSize=9, textColor=GRIS,
            alignment=TA_CENTER,
        ),
        "seccion": ParagraphStyle(
            "seccion", fontName="Helvetica-Bold", fontSize=11, textColor=ORO,
            spaceBefore=6, spaceAfter=4,
        ),
        "label": ParagraphStyle(
            "label", fontName="Helvetica-Bold", fontSize=9, textColor=NEGRO,
        ),
        "valor": ParagraphStyle(
            "valor", fontName="Helvetica", fontSize=9.5, textColor=NEGRO,
        ),
        "th": ParagraphStyle(
            "th", fontName="Helvetica-Bold", fontSize=8.5, textColor=BLANCO,
            alignment=TA_CENTER,
        ),
        "td": ParagraphStyle(
            "td", fontName="Helvetica", fontSize=8.5, textColor=NEGRO,
        ),
        "tdc": ParagraphStyle(
            "tdc", fontName="Helvetica", fontSize=8.5, textColor=NEGRO,
            alignment=TA_CENTER,
        ),
        "tdr": ParagraphStyle(
            "tdr", fontName="Helvetica", fontSize=8.5, textColor=NEGRO,
            alignment=TA_RIGHT,
        ),
        "total_label": ParagraphStyle(
            "total_label", fontName="Helvetica-Bold", fontSize=10,
            textColor=NEGRO, alignment=TA_RIGHT,
        ),
        "total_valor": ParagraphStyle(
            "total_valor", fontName="Helvetica-Bold", fontSize=12,
            textColor=ORO, alignment=TA_RIGHT,
        ),
        "nota": ParagraphStyle(
            "nota", fontName="Helvetica", fontSize=7.5, textColor=GRIS,
            alignment=TA_CENTER,
        ),
    }


def _celda(texto, estilo):
    return Paragraph(str(texto or ""), estilo)


def _cop(valor) -> str:
    try:
        v = float(valor or 0)
    except (TypeError, ValueError):
        v = 0
    return f"${v:,.0f} COP".replace(",", ".")


# ── Helpers PDF comunes ──────────────────────────────────────────


def _header(
    s: dict,
    titulo: str,
    periodo: str,
    inicio: date,
    fin: date,
    tecnico_nombre: str | None,
):
    """Devuelve una lista de flowables con el encabezado del reporte."""
    historia = []
    historia.append(Paragraph(f"<b>NEODOMUS</b>", s["titulo"]))
    historia.append(Paragraph("Reportes del Sistema", s["sub"]))
    historia.append(Spacer(1, 2 * mm))

    linea = Table([[""]], colWidths=[180 * mm], rowHeights=[1])
    linea.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), ORO),
        ("LINEBELOW", (0, 0), (-1, -1), 0, ORO),
    ]))
    historia.append(linea)
    historia.append(Spacer(1, 4 * mm))

    historia.append(Paragraph(f"<b>{titulo}</b>", s["seccion"]))
    filtro_texto = f"Periodo: {periodo.capitalize()} | Del {inicio} al {fin}"
    if tecnico_nombre:
        filtro_texto += f" | Técnico: {tecnico_nombre}"
    historia.append(Paragraph(filtro_texto, s["sub"]))
    historia.append(Spacer(1, 4 * mm))

    return historia


def _footer(s: dict, historia: list):
    """Agrega el pie de página al final."""
    historia.append(Spacer(1, 6 * mm))
    linea = Table([[""]], colWidths=[180 * mm], rowHeights=[1])
    linea.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), ORO),
    ]))
    historia.append(linea)
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M")
    historia.append(Paragraph(f"Generado el {ahora} - NEODOMUS", s["nota"]))


def _build_pdf(historia: list, titulo: str) -> io.BytesIO:
    """Construye el PDF en memoria y retorna un BytesIO."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"{titulo} - NEODOMUS",
        author="NEODOMUS",
    )
    doc.build(historia)
    buf.seek(0)
    return buf


# ── PDF: Reporte de Ventas ───────────────────────────────────────


def generar_ventas_pdf(
    resumen: dict,
    ventas_por_periodo: list[dict],
    periodo: str,
    inicio: date,
    fin: date,
    tecnico_nombre: str | None,
) -> io.BytesIO:
    s = _styles()
    historia = _header(s, "REPORTE DE VENTAS", periodo, inicio, fin, tecnico_nombre)

    # ── Resumen ──────────────────────────────────────────
    historia.append(Paragraph("<b>RESUMEN</b>", s["seccion"]))
    res_data = [
        [_celda("Total Pedidos", s["label"]), _celda(str(resumen["total_pedidos"]), s["valor"])],
        [_celda("Ventas Productos", s["label"]), _celda(_cop(resumen["total_ventas_pedidos"]), s["valor"])],
        [_celda("Ingresos Citas", s["label"]), _celda(_cop(resumen["total_ingresos_citas"]), s["valor"])],
    ]
    res_tabla = Table(res_data, colWidths=[60 * mm, 80 * mm])
    res_tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), FONDO_TABLA),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.5, GRIS),
    ]))
    historia.append(res_tabla)
    historia.append(Spacer(1, 2 * mm))

    # Total destacado
    total_data = [[
        _celda("TOTAL INGRESOS", s["total_label"]),
        _celda(_cop(resumen["total_ingresos"]), s["total_valor"]),
    ]]
    total_tabla = Table(total_data, colWidths=[60 * mm, 80 * mm])
    total_tabla.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    historia.append(total_tabla)
    historia.append(Spacer(1, 6 * mm))

    # ── Detalle por periodo ──────────────────────────────
    if ventas_por_periodo:
        historia.append(Paragraph("<b>DETALLE POR PERIODO</b>", s["seccion"]))
        headers = ["Periodo", "Pedidos", "Ventas Productos", "Ingresos Citas", "Total"]
        enc_row = [_celda(h, s["th"]) for h in headers]
        rows = [enc_row]
        for v in ventas_por_periodo:
            rows.append([
                _celda(str(v["periodo"]), s["tdc"]),
                _celda(str(v["pedidos"]), s["tdc"]),
                _celda(_cop(v["ventas_pedidos"]), s["tdr"]),
                _celda(_cop(v["ingresos_citas"]), s["tdr"]),
                _celda(_cop(v["total"]), s["tdr"]),
            ])

        col_w = [30 * mm, 22 * mm, 38 * mm, 38 * mm, 38 * mm]
        det_tabla = Table(rows, colWidths=col_w, repeatRows=1)
        det_tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ORO),
            ("TEXTCOLOR", (0, 0), (-1, 0), BLANCO),
            ("BACKGROUND", (0, 1), (-1, -1), FONDO_TABLA),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [FONDO_TABLA, BLANCO]),
            ("GRID", (0, 0), (-1, -1), 0.5, GRIS),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        historia.append(det_tabla)

    _footer(s, historia)
    return _build_pdf(historia, "Reporte de Ventas")


# ── PDF: Reporte de Citas ────────────────────────────────────────


def generar_citas_pdf(
    resumen: dict,
    citas_por_periodo: list[dict],
    periodo: str,
    inicio: date,
    fin: date,
    tecnico_nombre: str | None,
) -> io.BytesIO:
    s = _styles()
    historia = _header(s, "REPORTE DE CITAS", periodo, inicio, fin, tecnico_nombre)

    # ── Resumen ──────────────────────────────────────────
    historia.append(Paragraph("<b>RESUMEN</b>", s["seccion"]))
    pe = resumen["por_estado"]
    res_data = [
        [_celda("Total Citas", s["label"]), _celda(str(resumen["total_citas"]), s["valor"])],
        [_celda("Pendiente", s["label"]), _celda(str(pe.get("Pendiente", 0)), s["valor"])],
        [_celda("Confirmada", s["label"]), _celda(str(pe.get("Confirmada", 0)), s["valor"])],
        [_celda("Finalizada", s["label"]), _celda(str(pe.get("Finalizada", 0)), s["valor"])],
        [_celda("Cancelada", s["label"]), _celda(str(pe.get("Cancelada", 0)), s["valor"])],
        [_celda("Ingresos Totales", s["label"]), _celda(_cop(resumen["ingresos_total"]), s["valor"])],
    ]
    res_tabla = Table(res_data, colWidths=[60 * mm, 80 * mm])
    res_tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), FONDO_TABLA),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.5, GRIS),
    ]))
    historia.append(res_tabla)
    historia.append(Spacer(1, 6 * mm))

    # ── Detalle por periodo ──────────────────────────────
    if citas_por_periodo:
        historia.append(Paragraph("<b>DETALLE POR PERIODO</b>", s["seccion"]))
        headers = ["Periodo", "Total", "Pendiente", "Confirmada", "Finalizada", "Cancelada"]
        enc_row = [_celda(h, s["th"]) for h in headers]
        rows = [enc_row]
        for c in citas_por_periodo:
            rows.append([
                _celda(str(c["periodo"]), s["tdc"]),
                _celda(str(c["total"]), s["tdc"]),
                _celda(str(c.get("Pendiente", 0)), s["tdc"]),
                _celda(str(c.get("Confirmada", 0)), s["tdc"]),
                _celda(str(c.get("Finalizada", 0)), s["tdc"]),
                _celda(str(c.get("Cancelada", 0)), s["tdc"]),
            ])

        col_w = [30 * mm, 22 * mm, 26 * mm, 28 * mm, 28 * mm, 28 * mm]
        det_tabla = Table(rows, colWidths=col_w, repeatRows=1)
        det_tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ORO),
            ("TEXTCOLOR", (0, 0), (-1, 0), BLANCO),
            ("BACKGROUND", (0, 1), (-1, -1), FONDO_TABLA),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [FONDO_TABLA, BLANCO]),
            ("GRID", (0, 0), (-1, -1), 0.5, GRIS),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        historia.append(det_tabla)

    _footer(s, historia)
    return _build_pdf(historia, "Reporte de Citas")


# ── Excel helpers ────────────────────────────────────────────────

ORO_HEX = "caa24d"
FONDO_HEX = "f7f3ea"
GRIS_HEX = "6b6b6b"

_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill(start_color=ORO_HEX, end_color=ORO_HEX, fill_type="solid")
_label_font = Font(name="Calibri", bold=True, size=11)
_valor_font = Font(name="Calibri", size=11)
_total_font = Font(name="Calibri", bold=True, size=13, color=ORO_HEX)
_thin_border = Border(
    left=Side(style="thin", color=GRIS_HEX),
    right=Side(style="thin", color=GRIS_HEX),
    top=Side(style="thin", color=GRIS_HEX),
    bottom=Side(style="thin", color=GRIS_HEX),
)
_center = Alignment(horizontal="center", vertical="center")
_right = Alignment(horizontal="right", vertical="center")
_left = Alignment(horizontal="left", vertical="center")


def _auto_width(ws, cols: int):
    for col_idx in range(1, cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


# ── Excel: Reporte de Ventas ─────────────────────────────────────


def generar_ventas_excel(
    resumen: dict,
    ventas_por_periodo: list[dict],
    periodo: str,
    inicio: date,
    fin: date,
    tecnico_nombre: str | None,
) -> io.BytesIO:
    wb = Workbook()

    # ── Hoja Resumen ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.append(["REPORTE DE VENTAS"])
    ws1.merge_cells("A1:B1")
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color=ORO_HEX)

    ws1.append([f"Periodo: {periodo.capitalize()}"])
    ws1.append([f"Del {inicio} al {fin}"])
    if tecnico_nombre:
        ws1.append([f"Técnico: {tecnico_nombre}"])
    ws1.append([])

    headers_r = ["Concepto", "Valor"]
    ws1.append(headers_r)
    for cell in ws1[ws1.max_row]:
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = _thin_border
        cell.alignment = _center

    res_rows = [
        ("Total Pedidos", resumen["total_pedidos"]),
        ("Ventas Productos", resumen["total_ventas_pedidos"]),
        ("Ingresos Citas", resumen["total_ingresos_citas"]),
        ("TOTAL INGRESOS", resumen["total_ingresos"]),
    ]
    for label, val in res_rows:
        ws1.append([label, val])
        ws1.cell(row=ws1.max_row, column=1).font = _label_font
        ws1.cell(row=ws1.max_row, column=1).border = _thin_border
        ws1.cell(row=ws1.max_row, column=1).alignment = _left
        ws1.cell(row=ws1.max_row, column=2).font = _total_font if "TOTAL" in label else _valor_font
        ws1.cell(row=ws1.max_row, column=2).border = _thin_border
        ws1.cell(row=ws1.max_row, column=2).alignment = _right
        ws1.cell(row=ws1.max_row, column=2).number_format = '#,##0'

    _auto_width(ws1, 2)

    # ── Hoja Detalle ─────────────────────────────────────
    ws2 = wb.create_sheet("Detalle")
    headers_d = ["Periodo", "Pedidos", "Ventas Productos", "Ingresos Citas", "Total"]
    ws2.append(headers_d)
    for cell in ws2[1]:
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = _thin_border
        cell.alignment = _center

    for v in ventas_por_periodo:
        ws2.append([
            str(v["periodo"]),
            v["pedidos"],
            v["ventas_pedidos"],
            v["ingresos_citas"],
            v["total"],
        ])
        row_num = ws2.max_row
        ws2.cell(row=row_num, column=1).alignment = _center
        for col in range(2, 6):
            ws2.cell(row=row_num, column=col).number_format = '#,##0'
            ws2.cell(row=row_num, column=col).alignment = _right
        for col in range(1, 6):
            ws2.cell(row=row_num, column=col).border = _thin_border

    _auto_width(ws2, 5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Excel: Reporte de Citas ──────────────────────────────────────


def generar_citas_excel(
    resumen: dict,
    citas_por_periodo: list[dict],
    periodo: str,
    inicio: date,
    fin: date,
    tecnico_nombre: str | None,
) -> io.BytesIO:
    wb = Workbook()

    # ── Hoja Resumen ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.append(["REPORTE DE CITAS"])
    ws1.merge_cells("A1:B1")
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color=ORO_HEX)

    ws1.append([f"Periodo: {periodo.capitalize()}"])
    ws1.append([f"Del {inicio} al {fin}"])
    if tecnico_nombre:
        ws1.append([f"Técnico: {tecnico_nombre}"])
    ws1.append([])

    ws1.append(["Concepto", "Valor"])
    for cell in ws1[ws1.max_row]:
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = _thin_border
        cell.alignment = _center

    pe = resumen["por_estado"]
    res_rows = [
        ("Total Citas", resumen["total_citas"]),
        ("Pendiente", pe.get("Pendiente", 0)),
        ("Confirmada", pe.get("Confirmada", 0)),
        ("Finalizada", pe.get("Finalizada", 0)),
        ("Cancelada", pe.get("Cancelada", 0)),
        ("Ingresos Totales", resumen["ingresos_total"]),
    ]
    for label, val in res_rows:
        ws1.append([label, val])
        ws1.cell(row=ws1.max_row, column=1).font = _label_font
        ws1.cell(row=ws1.max_row, column=1).border = _thin_border
        ws1.cell(row=ws1.max_row, column=1).alignment = _left
        ws1.cell(row=ws1.max_row, column=2).font = _valor_font
        ws1.cell(row=ws1.max_row, column=2).border = _thin_border
        ws1.cell(row=ws1.max_row, column=2).alignment = _right
        ws1.cell(row=ws1.max_row, column=2).number_format = '#,##0'

    _auto_width(ws1, 2)

    # ── Hoja Detalle ─────────────────────────────────────
    ws2 = wb.create_sheet("Detalle")
    headers_d = ["Periodo", "Total", "Pendiente", "Confirmada", "Finalizada", "Cancelada"]
    ws2.append(headers_d)
    for cell in ws2[1]:
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = _thin_border
        cell.alignment = _center

    for c in citas_por_periodo:
        ws2.append([
            str(c["periodo"]),
            c["total"],
            c.get("Pendiente", 0),
            c.get("Confirmada", 0),
            c.get("Finalizada", 0),
            c.get("Cancelada", 0),
        ])
        row_num = ws2.max_row
        ws2.cell(row=row_num, column=1).alignment = _center
        for col in range(2, 7):
            ws2.cell(row=row_num, column=col).alignment = _center
        for col in range(1, 7):
            ws2.cell(row=row_num, column=col).border = _thin_border

    _auto_width(ws2, 6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
